import openpyxl as op
from openpyxl.styles import NamedStyle
import pandas as pd
import datetime as dt
from datetime import timezone

# Pegando data de hoje
data_hoje = dt.datetime.now()
# Criando a classe para fornecedores / pedidos / emails


class Fornecedor():
    def __init__(self, Nome, Email):
        self.Nome = Nome
        self.Email = Email
        self.TotalPedidos = []

    def inserirPedidos(self, incremento_pedidos):
        self.TotalPedidos.append(incremento_pedidos)

    def mostrarPedidos(self):
        print(self.TotalPedidos)

    def removerPedido(self, Valor):
        print("------------------------")
        self.Valor = Valor
        pedido_removido = self.TotalPedidos[Valor]
        self.TotalPedidos.pop(Valor)
        print("Pedidos após a remoção-----------------")
        print(self.TotalPedidos)
        print("------------------------")


Pedidos = pd.read_excel('EntregasPendentes10_07_2023.xlsx')

Pedidos = Pedidos[Pedidos['Situação'] != 'Envio pendente']

Pedidos = Pedidos[Pedidos['Nacionalidade'] == 'Brasil']

valoresRateio = ['MATERIA-PRIMA',
                 'MATERIA PRIMA INDUSTRIALIZAÇÃO', 'MATERIAL DE USO E CONSUMO']
Pedidos = Pedidos[Pedidos['Rateio'].isin(valoresRateio)]

print(Pedidos)

Pedidos.to_excel('TestePandas.xlsx')

""" # Carregando Planilha
ws = op.load_workbook('./EntregasPendentes10_07_2023.xlsx')
planilha_ativa = ws.active
ult_linha_planilha = planilha_ativa.max_row

# Vendo todos pedidos que estão atrasados
def pegandoPedidosAtrasados(database):
    i=0
    fill_atrasado = Pattern
    ultima_linha = database.max_row
    datas_antigas = [planilha_ativa.cell(row=i, column=2).value for i in range(2,ultima_linha):]
    
    if(planilha_ativa.cell(row=i, column=2)<data_hoje: planilha_ativa[f"B.{celula.row}"].fill = PatternFill(start_color)


# Limpando planilha
for celula in planilha_ativa["Q"]:
    if (celula.value == "Envio pendente"):
        linha_celula = celula.row
        planilha_ativa.delete_rows(linha_celula)

for celula in planilha_ativa["G"]:
    if (celula.value != "Brasil" and celula.value != "Nacionalidade"):
        linha_celula = celula.row
        planilha_ativa.delete_rows(linha_celula)

for celula in planilha_ativa['M']:
    if (celula.value != "MATERIA-PRIMA" and celula.value != "Rateio"):
        linha_celula = celula.row
        planilha_ativa.delete_rows(linha_celula)


ws.save("PedidoAtraso.xlsx") """

""" tabelapd = pd.read_excel("./PedidoAtraso.xlsx")

# Puxando fornecedores sem duplicatas
fornecedores = tabelapd.loc[:, ['Fornecedor']].drop_duplicates(
    subset="Fornecedor", keep="first").values.tolist()

pedidosTMF = tabelapd.loc[tabelapd['Fornecedor'] == 'TMF COMPONENTES ELETRO EL ETRONICOS LTDA', [
    "Neg.", "Data de entrega", "Cod.", "Material", "Faltam"]].reset_index()
pedidosTMF.pop(pedidosTMF.columns[0])

pedidosTMF.index += 1

pedidosTMF['Data de entrega'] = pd.to_datetime(
    pedidosTMF['Data de entrega'], format='%d/%m/%Y')

pedidosTMF['Data de entrega'] = pedidosTMF["Data de entrega"].dt.strftime(
    "%d/%m/%Y   ")

pedidosTMF.to_excel('PedidosTMF.xlsx') """
